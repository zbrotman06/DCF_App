"""
╔══════════════════════════════════════════════════════════════╗
║          PROFESSIONAL DCF VALUATION ENGINE                   ║
║   Real-time data · UFCF Build-up · Dual Methods · Sensitivity║
╚══════════════════════════════════════════════════════════════╝
Requirements:
    pip install streamlit yfinance pandas numpy openpyxl plotly requests
Run:
    streamlit run dcf_app.py
"""

import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import io
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="DCF Valuation Engine",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
NEW_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;1,400&family=Instrument+Serif:ital@0;1&family=Geist+Mono:wght@300;400;500;600&display=swap');
 
/* ── CSS Variables ── */
:root {
    --bg-base:     #050810;
    --bg-subtle:   #0a0e18;
    --bg-card:     #0d1117;
    --bg-muted:    #111827;
    --bg-overlay:  #161f2e;
    --bg-border:   #1a2535;
    --bg-divider:  #1e2d3d;
 
    --green:       #2ea043;
    --green-text:  #3fb950;
    --green-dim:   rgba(63,185,80,0.12);
    --green-border:rgba(63,185,80,0.25);
 
    --red:         #c93737;
    --red-text:    #f85149;
    --red-dim:     rgba(248,81,73,0.10);
    --red-border:  rgba(248,81,73,0.25);
 
    --blue:        #1158c7;
    --blue-text:   #58a6ff;
    --blue-dim:    rgba(88,166,255,0.10);
    --blue-border: rgba(88,166,255,0.22);
 
    --gold-text:   #d29922;
    --gold-dim:    rgba(210,153,34,0.10);
    --gold-border: rgba(210,153,34,0.25);
 
    --text-1: #e6edf3;
    --text-2: #8b949e;
    --text-3: #57606a;
    --text-4: #30363d;
    --text-5: #21262d;
 
    --font-display: 'DM Sans', sans-serif;
    --font-serif:   'Instrument Serif', Georgia, serif;
    --font-mono:    'Geist Mono', monospace;
}
 
/* ── Page base ── */
*, html, body, [class*="css"] { font-family: var(--font-display); box-sizing: border-box; }
.stApp {
    background-color: var(--bg-base);
    background-image:
        linear-gradient(rgba(26,37,53,0.25) 1px, transparent 1px),
        linear-gradient(90deg, rgba(26,37,53,0.25) 1px, transparent 1px);
    background-size: 48px 48px;
}
.block-container { padding: 1.5rem 2.5rem !important; max-width: 100% !important; }
#MainMenu, footer, header { visibility: hidden; }
 
/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #080c14 0%, var(--bg-base) 100%) !important;
    border-right: 1px solid var(--bg-border);
}
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span { color: var(--text-3) !important; font-size: 12px !important; }
[data-testid="stSidebar"] .stTextInput input {
    background: var(--bg-muted) !important;
    border: 1px solid var(--bg-divider) !important;
    color: var(--text-1) !important;
    border-radius: 8px !important;
    font-family: var(--font-mono) !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    letter-spacing: 1.5px !important;
    padding: 10px 14px !important;
}
[data-testid="stSidebar"] .stTextInput input:focus {
    border-color: var(--blue-text) !important;
    box-shadow: 0 0 0 3px rgba(88,166,255,0.08) !important;
}
 
/* ── Sidebar expanders ── */
[data-testid="stSidebar"] .streamlit-expanderHeader {
    font-family: var(--font-mono) !important;
    font-size: 10px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 1.8px !important;
    color: var(--text-4) !important;
    background: transparent !important;
    border: none !important;
    border-top: 1px solid var(--bg-border) !important;
    padding: 12px 4px !important;
}
[data-testid="stSidebar"] .streamlit-expanderHeader:hover {
    color: var(--text-3) !important;
}
 
/* ── Tabs ── */
[data-testid="stTabs"] [role="tablist"] {
    border-bottom: 1px solid var(--bg-border);
    gap: 0;
    background: transparent;
}
[data-testid="stTabs"] [role="tab"] {
    font-family: var(--font-mono) !important;
    font-size: 10px !important;
    font-weight: 500 !important;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    color: var(--text-4) !important;
    padding: 11px 18px !important;
    border: none !important;
    background: transparent !important;
    border-bottom: 2px solid transparent !important;
    transition: all 0.18s;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    color: var(--blue-text) !important;
    border-bottom: 2px solid var(--blue-text) !important;
}
[data-testid="stTabs"] [role="tab"]:hover { color: var(--text-2) !important; }
 
/* ── KPI Cards — two-tier hierarchy ── */
.kpi-grid {
    display: grid;
    grid-template-columns: 1.6fr 1.6fr 1fr 1fr 1fr 1fr;
    gap: 10px;
    margin: 18px 0;
    align-items: stretch;
}
.kpi-card {
    background: var(--bg-card);
    border: 1px solid var(--bg-border);
    border-radius: 12px;
    padding: 18px 20px;
    position: relative;
    overflow: hidden;
    transition: border-color 0.18s;
}
.kpi-card:hover { border-color: var(--bg-divider); }
.kpi-card.primary { padding: 22px 24px; }
.kpi-accent { position: absolute; top: 0; left: 0; width: 100%; height: 2px; }
.kpi-label {
    font-family: var(--font-mono);
    font-size: 9px;
    font-weight: 500;
    text-transform: uppercase;
    letter-spacing: 2px;
    color: var(--text-4);
    margin-bottom: 10px;
}
.kpi-value {
    font-family: var(--font-mono);
    font-variant-numeric: tabular-nums;
    font-weight: 600;
    color: var(--text-1);
    line-height: 1;
}
.kpi-card.primary .kpi-value { font-size: 28px; }
.kpi-card:not(.primary) .kpi-value { font-size: 19px; }
.kpi-sub {
    font-family: var(--font-mono);
    font-size: 10px;
    color: var(--text-4);
    margin-top: 7px;
    font-variant-numeric: tabular-nums;
}
.pos { color: var(--green-text); }
.neg { color: var(--red-text);   }
 
/* ── Hero ── */
.hero { padding: 6px 0 22px 0; }
.hero-ticker {
    font-family: var(--font-serif);
    font-size: 44px;
    font-weight: 400;
    letter-spacing: -1.5px;
    color: var(--text-1);
    line-height: 1;
}
.hero-name { font-size: 13px; color: var(--text-4); margin: 5px 0 14px; font-weight: 400; }
.hero-price { font-family: var(--font-mono); font-size: 13px; color: var(--text-3); font-variant-numeric: tabular-nums; }
 
/* ── Section label ── */
.sec-label {
    font-family: var(--font-mono);
    font-size: 9px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 2.5px;
    color: var(--text-5);
    padding: 20px 0 10px 0;
    border-bottom: 1px solid rgba(255,255,255,0.03);
    margin-bottom: 14px;
}
 
/* ── Financial tables ── */
.fin-table {
    width: 100%;
    border-collapse: collapse;
    font-family: var(--font-mono);
    font-size: 11.5px;
    font-variant-numeric: tabular-nums;
}
.fin-table th {
    background: var(--bg-base);
    color: var(--text-4);
    font-size: 9px;
    text-transform: uppercase;
    letter-spacing: 1.4px;
    padding: 10px 14px;
    text-align: right;
    border-bottom: 1px solid var(--bg-border);
    font-weight: 600;
    position: sticky;
    top: 0;
    z-index: 1;
}
.fin-table th:first-child { text-align: left; padding-left: 6px; }
.fin-table td {
    padding: 7px 14px;
    text-align: right;
    border-bottom: 1px solid rgba(255,255,255,0.025);
    color: var(--text-2);
    font-variant-numeric: tabular-nums;   /* ← CRITICAL: aligns decimal points */
}
.fin-table td:first-child { text-align: left; padding-left: 6px; color: var(--text-3); }
.fin-table td.negative   { color: var(--red-text) !important; }
.fin-table td.positive   { color: var(--green-text) !important; }
.fin-table tr:hover td   { background: rgba(255,255,255,0.015); }
.fin-table .tot { border-top: 1px solid var(--bg-border); }
.fin-table .tot td {
    color: var(--text-1) !important;
    font-weight: 600;
    background: var(--bg-muted);
    border-bottom: 2px solid var(--green);
}
.fin-table .sub td { color: var(--text-4); font-size: 10.5px; }
 
/* ── Badges ── */
.badge {
    display: inline-flex; align-items: center; gap: 5px;
    border-radius: 6px; padding: 4px 10px;
    font-family: var(--font-mono); font-size: 11px; font-weight: 600;
}
.badge-up  { background: var(--green-dim); border: 1px solid var(--green-border); color: var(--green-text); }
.badge-dn  { background: var(--red-dim);   border: 1px solid var(--red-border);   color: var(--red-text); }
.badge-neu { background: var(--blue-dim);  border: 1px solid var(--blue-border);  color: var(--blue-text); }
 
/* ── Info / warning boxes ── */
.info-box {
    background: var(--blue-dim); border: 1px solid var(--blue-border);
    border-radius: 8px; padding: 12px 16px;
    font-size: 12px; color: var(--text-2);
    margin: 10px 0; font-family: var(--font-mono); line-height: 1.7;
}
.warn-box {
    background: var(--gold-dim); border: 1px solid var(--gold-border);
    border-left: 3px solid var(--gold-text);
    border-radius: 0 8px 8px 0; padding: 12px 16px;
    font-size: 12px; color: var(--gold-text);
    margin: 10px 0; font-family: var(--font-mono); line-height: 1.7;
}
.err-box {
    background: var(--red-dim); border: 1px solid var(--red-border);
    border-left: 3px solid var(--red-text);
    border-radius: 0 8px 8px 0; padding: 12px 16px;
    font-size: 12px; color: var(--red-text);
    margin: 10px 0; font-family: var(--font-mono); line-height: 1.7;
}
 
/* ── Scenario pills ── */
.scenario-bar { display: flex; gap: 8px; margin: 10px 0 18px; }
.sc-pill {
    border-radius: 6px; padding: 5px 14px;
    font-family: var(--font-mono); font-size: 11px; font-weight: 600;
    border: 1px solid; cursor: default; letter-spacing: 0.3px;
}
.sc-base { background: var(--blue-dim);  border-color: var(--blue-border);  color: var(--blue-text); }
.sc-bull { background: var(--green-dim); border-color: var(--green-border); color: var(--green-text); }
.sc-bear { background: var(--red-dim);   border-color: var(--red-border);   color: var(--red-text); }
 
/* ── Sidebar section header ── */
.sb-section {
    font-family: var(--font-mono); font-size: 9px; font-weight: 600;
    text-transform: uppercase; letter-spacing: 2px; color: var(--text-5);
    border-top: 1px solid var(--bg-border); padding: 14px 0 6px; margin-top: 6px;
}
 
/* ── Buttons ── */
.stDownloadButton button {
    background: linear-gradient(135deg, #1a7f37, #238636) !important;
    border: 1px solid #2ea043 !important; color: #fff !important;
    border-radius: 8px !important; font-family: var(--font-mono) !important;
    font-size: 12px !important; font-weight: 600 !important;
    letter-spacing: 0.5px !important; padding: 10px 20px !important;
    width: 100%; transition: all 0.18s !important;
}
.stButton button {
    background: linear-gradient(135deg, #0f3d8a, var(--blue)) !important;
    border: 1px solid var(--blue-text) !important; color: #fff !important;
    border-radius: 8px !important; font-family: var(--font-mono) !important;
    font-size: 12px !important; font-weight: 600 !important;
    padding: 11px !important; width: 100% !important;
}
 
/* ── Divider ── */
.divider { border: none; border-top: 1px solid var(--bg-border); margin: 22px 0; }
 
/* ── Footer ── */
.footer {
    font-family: var(--font-mono); font-size: 10px; color: var(--text-5);
    text-align: center; padding: 22px 0; letter-spacing: 1.2px;
}
 
/* ── st.metric override ── */
[data-testid="stMetric"] {
    background: var(--bg-card);
    border: 1px solid var(--bg-border);
    border-radius: 10px;
    padding: 14px 18px;
}
[data-testid="stMetricLabel"] { font-family: var(--font-mono) !important; font-size: 10px !important; color: var(--text-4) !important; text-transform: uppercase; letter-spacing: 1.5px; }
[data-testid="stMetricValue"] { font-family: var(--font-mono) !important; font-variant-numeric: tabular-nums; color: var(--text-1) !important; }
[data-testid="stMetricDelta"] { font-family: var(--font-mono) !important; font-variant-numeric: tabular-nums; }
 
/* ── Plotly container ── */
.js-plotly-plot .plotly .main-svg { background: transparent !important; }
 
/* ── Home nav button (hidden, triggered by JS) ── */
.home-nav-btn { display: none; }
</style>
"""
st.markdown(NEW_CSS, unsafe_allow_html=True)

# ═══════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════

def fmt(val, style="$M", decimals=1):
    """Format a number for display."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "N/A"
    if style == "$M":
        return f"${val/1e6:,.{decimals}f}M"
    elif style == "$B":
        return f"${val/1e9:,.{decimals}f}B"
    elif style == "%":
        return f"{val*100:.{decimals}f}%"
    elif style == "x":
        return f"{val:.{decimals}f}x"
    elif style == "$":
        return f"${val:,.{decimals}f}"
    return str(val)


@st.cache_data(ttl=300, show_spinner=False)
def fetch_data(ticker: str):
    """Pull all financial data from yfinance."""
    import time, random
    time.sleep(random.uniform(1, 2))
    t = yf.Ticker(ticker)
    try:
        info = t.info or {}
    except Exception:
        info = {}

    # Pull statements — try annual first
    try:
        inc = t.financials          # income statement (annual)
        bal = t.balance_sheet       # balance sheet (annual)
        cf  = t.cashflow            # cash flow (annual)
        inc_q = t.quarterly_financials
    except Exception:
        return None, "Could not retrieve financial statements."

    if inc is None or inc.empty:
        return None, f"No financial data found for '{ticker}'."

    hist = t.history(period="5y")

    return {
        "info": info,
        "income": inc,
        "balance": bal,
        "cashflow": cf,
        "income_q": inc_q,
        "history": hist,
    }, None


def safe_get(df, *keys):
    """Safely retrieve a row from a DataFrame by trying multiple key names."""
    if df is None or df.empty:
        return pd.Series(dtype=float)
    for k in keys:
        if k in df.index:
            return pd.to_numeric(df.loc[k], errors="coerce")
    return pd.Series(dtype=float)


def build_historical_ufcf(data):
    """Build UFCF bridge from raw financial data."""
    inc = data["income"]
    bal = data["balance"]
    cf  = data["cashflow"]

    cols = inc.columns  # Annual periods

    ebit       = safe_get(inc, "EBIT", "Operating Income", "Ebit")
    tax_prov   = safe_get(inc, "Tax Provision", "Income Tax Expense")
    pretax     = safe_get(inc, "Pretax Income", "Income Before Tax")
    revenue    = safe_get(inc, "Total Revenue", "Revenue")
    cogs       = safe_get(inc, "Cost Of Revenue", "Cost of Goods Sold")
    gross      = safe_get(inc, "Gross Profit")
    sga        = safe_get(inc, "Selling General Administrative", "Selling General And Administrative")
    da         = safe_get(cf,  "Depreciation", "Depreciation And Amortization",
                               "Depreciation Amortization Depletion")
    capex      = safe_get(cf,  "Capital Expenditures", "Capital Expenditure",
                               "Purchase Of Ppe")
    chg_nwc    = safe_get(cf,  "Change In Working Capital", "Changes In Working Capital",
                               "Change In Other Working Capital")

    # Tax rate: use effective rate, clamp 10-40%
    eff_tax = (tax_prov / pretax.replace(0, np.nan)).clip(0.10, 0.40)
    eff_tax = eff_tax.fillna(0.21)

    # NOPAT = EBIT × (1 - t)
    nopat = ebit * (1 - eff_tax)

    # D&A (add back positive)
    da_pos = da.abs()

    # CapEx (make negative)
    capex_neg = -capex.abs()

    # UFCF = NOPAT + D&A + ΔWC + CapEx
    ufcf = nopat + da_pos + chg_nwc + capex_neg

    result = pd.DataFrame({
        "Revenue":      revenue,
        "Gross Profit": gross,
        "EBIT":         ebit,
        "Eff. Tax Rate":eff_tax,
        "NOPAT":        nopat,
        "D&A":          da_pos,
        "ΔWorking Cap": chg_nwc,
        "CapEx":        capex_neg,
        "UFCF":         ufcf,
    }).T

    result.columns = [str(c)[:10] for c in cols]
    result = result.apply(pd.to_numeric, errors="coerce")
    return result


def project_ufcf(hist_df, rev_growth_rates, ebit_margin, da_pct, capex_pct, nwc_pct, tax_rate, n_years=5):
    """
    Project UFCF for n_years.
    rev_growth_rates: list of annual revenue growth rates (length = n_years)
    """
    # Base year values from most recent column
    try:
        base_rev   = float(hist_df.loc["Revenue"].iloc[0])
        base_capex = float(hist_df.loc["CapEx"].iloc[0])
    except Exception:
        base_rev = 1e9
        base_capex = -0.05 * base_rev

    rows = []
    rev = base_rev
    prev_rev = rev

    for i, g in enumerate(rev_growth_rates):
        rev = prev_rev * (1 + g)
        ebit = rev * ebit_margin
        nopat = ebit * (1 - tax_rate)
        da = rev * da_pct
        capex = -abs(rev * capex_pct)
        d_nwc = -(rev - prev_rev) * nwc_pct   # increase in NWC = use of cash
        ufcf = nopat + da + d_nwc + capex

        rows.append({
            "Year":          f"Y+{i+1}",
            "Revenue":       rev,
            "Rev Growth":    g,
            "EBIT":          ebit,
            "EBIT Margin":   ebit_margin,
            "Tax Rate":      tax_rate,
            "NOPAT":         nopat,
            "D&A":           da,
            "D&A % Rev":     da_pct,
            "ΔWorking Cap":  d_nwc,
            "NWC % Rev":     nwc_pct,
            "CapEx":         capex,
            "CapEx % Rev":   capex_pct,
            "UFCF":          ufcf,
        })
        prev_rev = rev

    return pd.DataFrame(rows).set_index("Year")


def dcf_valuation(proj_df, wacc, tgr, exit_ebitda_mult, net_debt, shares_out):
    """Calculate intrinsic value via both terminal value methods."""
    ufcf_series = proj_df["UFCF"].values
    n = len(ufcf_series)

    # Discount factors
    disc = np.array([(1 / (1 + wacc) ** t) for t in range(1, n + 1)])
    pv_fcf = ufcf_series * disc
    sum_pv_fcf = pv_fcf.sum()

    # Terminal value — Gordon Growth Model
    tv_ggm = ufcf_series[-1] * (1 + tgr) / (wacc - tgr)
    pv_tv_ggm = tv_ggm / (1 + wacc) ** n

    # Terminal value — Exit Multiple (EV/EBITDA)
    terminal_ebitda = proj_df["EBIT"].iloc[-1]  # use EBIT as proxy (no D&A add-back here)
    tv_em = terminal_ebitda * exit_ebitda_mult
    pv_tv_em = tv_em / (1 + wacc) ** n

    def equity_val(pv_tv):
        ev = sum_pv_fcf + pv_tv
        eq = ev - net_debt
        price = eq / shares_out if shares_out > 0 else np.nan
        return ev, eq, price

    ev_ggm, eq_ggm, price_ggm = equity_val(pv_tv_ggm)
    ev_em,  eq_em,  price_em  = equity_val(pv_tv_em)

    return {
        "pv_fcf":       pv_fcf,
        "sum_pv_fcf":   sum_pv_fcf,
        # GGM
        "tv_ggm":       tv_ggm,
        "pv_tv_ggm":    pv_tv_ggm,
        "ev_ggm":       ev_ggm,
        "eq_ggm":       eq_ggm,
        "price_ggm":    price_ggm,
        # Exit Multiple
        "tv_em":        tv_em,
        "pv_tv_em":     pv_tv_em,
        "ev_em":        ev_em,
        "eq_em":        eq_em,
        "price_em":     price_em,
    }


def sensitivity_table(proj_df, base_wacc, base_tgr, base_mult, net_debt, shares, method="ggm"):
    """Build 6×6 sensitivity table varying WACC and TGR (or exit multiple)."""
    wacc_range = np.linspace(base_wacc - 0.03, base_wacc + 0.03, 7)
    if method == "ggm":
        param_range = np.linspace(max(base_tgr - 0.02, 0.005), base_tgr + 0.02, 7)
        param_label = "Terminal Growth Rate"
    else:
        param_range = np.linspace(max(base_mult - 4, 3), base_mult + 4, 7)
        param_label = "Exit EV/EBITDA"

    table = pd.DataFrame(index=[f"{w:.1%}" for w in wacc_range],
                         columns=[f"{p:.1%}" if method=="ggm" else f"{p:.1f}x" for p in param_range])

    for w in wacc_range:
        for p in param_range:
            if method == "ggm":
                if w <= p:
                    v = np.nan
                else:
                    r = dcf_valuation(proj_df, w, p, base_mult, net_debt, shares)
                    v = r["price_ggm"]
            else:
                r = dcf_valuation(proj_df, w, base_tgr, p, net_debt, shares)
                v = r["price_em"]
            table.loc[f"{w:.1%}", f"{p:.1%}" if method=="ggm" else f"{p:.1f}x"] = round(v, 2) if not np.isnan(v) else np.nan

    table.index.name = f"WACC \\ {param_label}"
    return table.astype(float)


def build_wacc(data, cost_of_equity, debt_weight=None):
    """Estimate WACC from balance sheet data."""
    info = data["info"]
    bal  = data["balance"]

    # Cost of debt: interest expense / total debt
    try:
        int_exp = abs(float(safe_get(data["income"],
                    "Interest Expense", "Interest Expense Non Operating").iloc[0]))
        total_debt_bs = float(safe_get(bal,
                    "Total Debt", "Long Term Debt And Capital Lease Obligation",
                    "Long Term Debt").iloc[0])
        cod = int_exp / total_debt_bs if total_debt_bs > 0 else 0.05
        cod = np.clip(cod, 0.02, 0.15)
    except Exception:
        cod = 0.05

    # Market cap
    mkt_cap = info.get("marketCap", 0) or 0

    # Weights
    total_cap = mkt_cap + (total_debt_bs if 'total_debt_bs' in dir() else 0)
    if total_cap > 0 and debt_weight is None:
        ew = mkt_cap / total_cap
        dw = 1 - ew
    elif debt_weight is not None:
        dw = debt_weight
        ew = 1 - dw
    else:
        ew, dw = 0.8, 0.2

    tax_rate = 0.21  # statutory
    wacc = ew * cost_of_equity + dw * cod * (1 - tax_rate)
    return wacc, cod, ew, dw


def export_excel(ticker, info, hist_df, proj_df, valuation, sensitivity_ggm, sensitivity_em,
                  wacc, tgr, exit_mult, net_debt, shares, current_price):
    """Build a multi-sheet Excel workbook and return bytes."""
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        wb = writer.book

        # ── Sheet 1: Summary ──────────────────────────
        summary_data = {
            "Metric": [
                "Company", "Ticker", "Current Price", "Shares Outstanding",
                "Net Debt", "WACC", "Terminal Growth Rate", "Exit EV/EBITDA Multiple",
                "─── GGM Method ───",
                "PV of FCFs", "PV of Terminal Value (GGM)", "Enterprise Value (GGM)",
                "Equity Value (GGM)", "Implied Share Price (GGM)",
                "Upside / (Downside) — GGM",
                "─── Exit Multiple Method ───",
                "PV of Terminal Value (Exit Mult)", "Enterprise Value (Exit Mult)",
                "Equity Value (Exit Mult)", "Implied Share Price (Exit Mult)",
                "Upside / (Downside) — Exit Mult",
            ],
            "Value": [
                info.get("longName", ticker), ticker,
                f"${current_price:.2f}",
                f"{shares/1e6:.1f}M",
                f"${net_debt/1e9:.2f}B",
                f"{wacc:.2%}",
                f"{tgr:.2%}",
                f"{exit_mult:.1f}x",
                "",
                fmt(valuation["sum_pv_fcf"], "$B"),
                fmt(valuation["pv_tv_ggm"], "$B"),
                fmt(valuation["ev_ggm"], "$B"),
                fmt(valuation["eq_ggm"], "$B"),
                f"${valuation['price_ggm']:.2f}",
                f"{(valuation['price_ggm']/current_price - 1):.1%}" if current_price else "N/A",
                "",
                fmt(valuation["pv_tv_em"], "$B"),
                fmt(valuation["ev_em"], "$B"),
                fmt(valuation["eq_em"], "$B"),
                f"${valuation['price_em']:.2f}",
                f"{(valuation['price_em']/current_price - 1):.1%}" if current_price else "N/A",
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        # ── Sheet 2: Historical UFCF ──────────────────
        hist_fmt = hist_df.copy() / 1e6
        hist_fmt.to_excel(writer, sheet_name="Historical UFCF")

        # ── Sheet 3: Projected UFCF ──────────────────
        proj_fmt = proj_df.copy()
        dollar_cols = ["Revenue","EBIT","NOPAT","D&A","ΔWorking Cap","CapEx","UFCF"]
        pct_cols    = ["Rev Growth","EBIT Margin","Tax Rate","D&A % Rev","NWC % Rev","CapEx % Rev"]
        for c in dollar_cols:
            if c in proj_fmt.columns:
                proj_fmt[c] = proj_fmt[c] / 1e6
        proj_fmt.to_excel(writer, sheet_name="Projected UFCF (USD M)")

        # ── Sheet 4: DCF Bridge ──────────────────────
        bridge = pd.DataFrame({
            "Component": ["PV of FCFs (Y1–Y" + str(len(proj_df)) + ")",
                          "PV of Terminal Value (GGM)", "Enterprise Value (GGM)",
                          "Less: Net Debt", "Equity Value (GGM)",
                          "Shares Outstanding", "Implied Price (GGM)",
                          "───",
                          "PV of Terminal Value (Exit Mult)", "Enterprise Value (Exit Mult)",
                          "Less: Net Debt", "Equity Value (Exit Mult)",
                          "Shares Outstanding", "Implied Price (Exit Mult)"],
            "USD ($B)": [
                valuation["sum_pv_fcf"]/1e9,
                valuation["pv_tv_ggm"]/1e9,
                valuation["ev_ggm"]/1e9,
                -net_debt/1e9,
                valuation["eq_ggm"]/1e9,
                shares/1e9,
                valuation["price_ggm"],
                None,
                valuation["pv_tv_em"]/1e9,
                valuation["ev_em"]/1e9,
                -net_debt/1e9,
                valuation["eq_em"]/1e9,
                shares/1e9,
                valuation["price_em"],
            ]
        })
        bridge.to_excel(writer, sheet_name="DCF Bridge", index=False)

        # ── Sheet 5: Sensitivity — GGM ───────────────
        sensitivity_ggm.to_excel(writer, sheet_name="Sensitivity — GGM")

        # ── Sheet 6: Sensitivity — Exit Mult ─────────
        sensitivity_em.to_excel(writer, sheet_name="Sensitivity — Exit Mult")

        # Style the workbook minimally
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter
        header_fill = PatternFill("solid", fgColor="1C2128")
        header_font = Font(bold=True, color="E6EDF3", name="Consolas", size=10)
        for shname in writer.sheets:
            ws = writer.sheets[shname]
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════
# PLOTTING FUNCTIONS
# ═══════════════════════════════════════════

CHART_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Geist Mono, JetBrains Mono, monospace", color="#57606a", size=10),
    margin=dict(l=10, r=10, t=38, b=10),
    legend=dict(
        bgcolor="rgba(0,0,0,0)",
        font=dict(size=10),
        orientation="h",
        yanchor="bottom", y=1.02, xanchor="right", x=1,
        bordercolor="rgba(0,0,0,0)",
    ),
    xaxis=dict(
        gridcolor="rgba(0,0,0,0)",    # no x gridlines
        linecolor="#1a2535",
        tickfont=dict(size=10),
        zeroline=False,
        showgrid=False,
    ),
    yaxis=dict(
        gridcolor="#111827",           # very subtle horizontal only
        gridwidth=1,
        linecolor="#1a2535",
        tickfont=dict(size=10),
        zeroline=False,
    ),
    hoverlabel=dict(
        bgcolor="#161f2e",
        bordercolor="#1a2535",
        font=dict(family="Geist Mono, monospace", size=11, color="#e6edf3"),
    ),
)

CHART_CONFIG = {"displayModeBar": False, "staticPlot": False, "responsive": True}

def plot_ufcf_history(hist_df):
    cols = hist_df.columns.tolist()
    rev  = hist_df.loc["Revenue"] / 1e9 if "Revenue" in hist_df.index else []
    ufcf = hist_df.loc["UFCF"] / 1e9 if "UFCF" in hist_df.index else []
    ebit = hist_df.loc["EBIT"] / 1e9 if "EBIT" in hist_df.index else []

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(x=cols, y=rev, name="Revenue", marker_color="#1f6feb", opacity=0.7))
    fig.add_trace(go.Bar(x=cols, y=ebit, name="EBIT", marker_color="#388bfd", opacity=0.8))
    fig.add_trace(go.Scatter(x=cols, y=ufcf, name="UFCF", mode="lines+markers",
                             line=dict(color="#3fb950", width=2.5),
                             marker=dict(size=7, symbol="circle")), secondary_y=True)
    fig.update_layout(title="Historical Financials ($B)", barmode="group", **CHART_LAYOUT)
    fig.update_yaxes(title_text="$B (Bars)", secondary_y=False, gridcolor="#21262d")
    fig.update_yaxes(title_text="UFCF $B", secondary_y=True, gridcolor="#21262d")
    return fig


def plot_projected_fcf(proj_df, pv_fcf):
    years = proj_df.index.tolist()
    ufcf  = proj_df["UFCF"].values / 1e9
    rev   = proj_df["Revenue"].values / 1e9
    pv    = pv_fcf / 1e9

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(go.Bar(x=years, y=rev, name="Projected Revenue", marker_color="#388bfd", opacity=0.5))
    fig.add_trace(go.Bar(x=years, y=ufcf, name="Projected UFCF", marker_color="#3fb950", opacity=0.9))
    fig.add_trace(go.Scatter(x=years, y=pv, name="PV of FCF", mode="lines+markers",
                             line=dict(color="#d29922", width=2, dash="dot"),
                             marker=dict(size=6)), secondary_y=True)
    fig.update_layout(title="Projected UFCF ($B)", barmode="group", **CHART_LAYOUT)
    return fig


def plot_valuation_bridge(val, net_debt, shares, method="ggm"):
    pv_fcf_b  = val["sum_pv_fcf"] / 1e9
    if method == "ggm":
        pv_tv_b = val["pv_tv_ggm"] / 1e9
        ev_b    = val["ev_ggm"] / 1e9
    else:
        pv_tv_b = val["pv_tv_em"] / 1e9
        ev_b    = val["ev_em"] / 1e9
    nd_b = net_debt / 1e9

    fig = go.Figure(go.Waterfall(
        name="Bridge",
        orientation="v",
        measure=["relative", "relative", "total", "relative", "total"],
        x=["PV FCFs", "PV Terminal\nValue", "Enterprise Value", "Less:\nNet Debt", "Equity Value"],
        y=[pv_fcf_b, pv_tv_b, 0, -nd_b, 0],
        connector=dict(line=dict(color="#30363d")),
        decreasing=dict(marker=dict(color="#f85149")),
        increasing=dict(marker=dict(color="#3fb950")),
        totals=dict(marker=dict(color="#388bfd")),
        text=[fmt(v*1e9,"$B") for v in [pv_fcf_b, pv_tv_b, ev_b, -nd_b,
                                          ev_b - nd_b]],
        textposition="outside",
    ))
    label = "GGM" if method == "ggm" else "Exit Multiple"
    fig.update_layout(title=f"Valuation Bridge — {label} ($B)", **CHART_LAYOUT)
    return fig


def plot_sensitivity_heatmap(sens_df, current_price):
    """Plotly heatmap for sensitivity table with current price overlay."""
    z = sens_df.values.astype(float)
    # Color: green above current, red below
    cmax = np.nanmax(z)
    cmin = np.nanmin(z)

    colorscale = [
        [0.0, "#da3633"],
        [0.4, "#7d8590"],
        [1.0, "#3fb950"],
    ]
    midpoint = max((current_price - cmin) / (cmax - cmin), 0) if (cmax > cmin) else 0.5

    fig = go.Figure(go.Heatmap(
        z=z,
        x=sens_df.columns.tolist(),
        y=sens_df.index.tolist(),
        colorscale=colorscale,
        zmid=current_price,
        text=[[f"${v:.2f}" if not np.isnan(v) else "N/A" for v in row] for row in z],
        texttemplate="%{text}",
        textfont=dict(size=10, family="IBM Plex Mono"),
        hovertemplate="WACC: %{y}<br>Param: %{x}<br>Price: %{text}<extra></extra>",
        showscale=True,
        colorbar=dict(tickfont=dict(family="IBM Plex Mono", size=9), title="$/share"),
    ))
    fig.update_layout(title="Sensitivity: Implied Share Price", **CHART_LAYOUT,
                      height=380)
    return fig


def plot_price_history_with_range(hist, price_ggm, price_em, current_price):
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=hist.index, y=hist["Close"],
        name="Share Price", line=dict(color="#388bfd", width=1.5),
        fill="tozeroy", fillcolor="rgba(56,139,253,0.08)"
    ))
    fig.add_hline(y=price_ggm, line_dash="dash", line_color="#3fb950",
                  annotation_text=f"GGM: ${price_ggm:.2f}",
                  annotation_font=dict(color="#3fb950", family="IBM Plex Mono", size=11))
    fig.add_hline(y=price_em, line_dash="dash", line_color="#d29922",
                  annotation_text=f"Exit Mult: ${price_em:.2f}",
                  annotation_font=dict(color="#d29922", family="IBM Plex Mono", size=11))
    fig.add_hline(y=current_price, line_dash="dot", line_color="#7d8590",
                  annotation_text=f"Current: ${current_price:.2f}",
                  annotation_font=dict(color="#7d8590", family="IBM Plex Mono", size=11))
    fig.update_layout(title="5-Year Price History vs. DCF Targets", height=300, **CHART_LAYOUT)
    return fig


# ═══════════════════════════════════════════
# SIDEBAR — INPUTS
# ═══════════════════════════════════════════

with st.sidebar:
    st.markdown("""
    <div style='padding:16px 0 6px 0;'>
      <div style='font-family:"Geist Mono",monospace;font-size:16px;font-weight:600;
           color:#e6edf3;letter-spacing:-0.3px;'>◈ DCF Engine</div>
      <div style='font-size:9px;color:#21262d;font-family:"Geist Mono",monospace;
           text-transform:uppercase;letter-spacing:2.5px;margin-top:4px;'>
           Professional Valuation Model</div>
    </div>""", unsafe_allow_html=True)
 
    # ── Company (always visible, prominent) ───────────────────
    st.markdown("<div style='margin:4px 0 2px;'></div>", unsafe_allow_html=True)
    ticker_input = st.text_input("", value="AAPL", max_chars=10,
                                  placeholder="Ticker — e.g. AAPL").upper().strip()
 
    # ── Scenario (always visible) ─────────────────────────────
    scenario_mode = st.selectbox("Scenario", ["Base","Bull","Bear"], index=0,
                                  label_visibility="collapsed")
 
    # ── Revenue Growth (collapsed by default) ────────────────
    with st.expander("Revenue & Growth", expanded=False):
        n_years = st.slider("Forecast Years", 3, 10, 5)
        uniform  = st.checkbox("Uniform growth rate", value=True)
        if uniform:
            base_g = st.slider("Base Growth",  -0.10, 0.40, 0.08, 0.005, format="%.1f%%")
            bull_g = st.slider("Bull Growth",  -0.10, 0.50, 0.14, 0.005, format="%.1f%%")
            bear_g = st.slider("Bear Growth",  -0.15, 0.30, 0.03, 0.005, format="%.1f%%")
            rev_growths_map = {
                "Base": [base_g]*n_years,
                "Bull": [bull_g]*n_years,
                "Bear": [bear_g]*n_years,
            }
        else:
            base_g = st.slider("Base Growth", -0.10, 0.40, 0.08, 0.005, format="%.1f%%")
            rev_growths_map = {"Base":[base_g]*n_years,"Bull":[],"Bear":[]}
            for i in range(n_years):
                rev_growths_map["Bull"].append(
                    st.slider(f"Bull Y{i+1}", -0.10, 0.50, max(0.18-i*0.02,0.05), 0.005, format="%.1f%%"))
                rev_growths_map["Bear"].append(
                    st.slider(f"Bear Y{i+1}", -0.15, 0.30, max(0.05-i*0.01,-0.02), 0.005, format="%.1f%%"))
        rev_growths = rev_growths_map[scenario_mode]
 
    # ── Margins & Reinvestment ────────────────────────────────
    with st.expander("Margins & Reinvestment", expanded=False):
        ebit_margin = st.slider("EBIT Margin",        0.0,  0.50, 0.20, 0.005, format="%.1f%%")
        da_pct      = st.slider("D&A % of Revenue",   0.01, 0.20, 0.05, 0.005, format="%.1f%%")
        capex_pct   = st.slider("CapEx % of Revenue", 0.01, 0.25, 0.06, 0.005, format="%.1f%%")
        nwc_pct     = st.slider("NWC Δ % of Rev Δ",   0.0,  0.30, 0.05, 0.005, format="%.1f%%")
        tax_rate    = st.slider("Tax Rate",            0.10, 0.40, 0.21, 0.01,  format="%.0f%%")
        sbc_pct     = st.slider("SBC % of Revenue",   0.0,  0.20, 0.0,  0.005, format="%.1f%%",
                                 help="Stock-based compensation. Material for tech companies.")
 
    # ── Cost of Capital ───────────────────────────────────────
    with st.expander("Cost of Capital", expanded=False):
        manual_wacc = st.checkbox("Manual WACC override")
        if manual_wacc:
            wacc_man = st.slider("WACC (%)", 4.0, 20.0, 9.0, 0.25, format="%.2f%%") / 100
            rf_rate = 0.043; erp = 0.055
        else:
            rf_rate      = st.slider("Risk-Free Rate",      0.02, 0.07, 0.043, 0.001, format="%.2f%%")
            erp          = st.slider("Equity Risk Premium", 0.03, 0.09, 0.055, 0.005, format="%.2f%%")
            beta_ov= st.checkbox("Override Beta")
            beta     = st.slider("Beta", 0.3, 2.5, 1.1, 0.05) if beta_ov else None
        mid_year = st.checkbox("Mid-year convention",
                                help="Discounts at t-0.5. More accurate, ~5% price uplift.")
        use_adj  = st.checkbox("Use SBC-adjusted UFCF", value=sbc_pct > 0)
 
    # ── Terminal Value ────────────────────────────────────────
    with st.expander("Terminal Value", expanded=False):
        tgr       = st.slider("Terminal Growth Rate", 0.005, 0.05, 0.025, 0.005, format="%.2f%%")
        exit_mult = st.slider("Exit EV/EBITDA",       5.0,  30.0,  12.0,  0.5,   format="%.1fx")
 
    # ── Comps ─────────────────────────────────────────────────
    with st.expander("Comparable Companies", expanded=False):
        run_comps   = st.checkbox("Fetch peer data", value=False)
        comps_input = st.text_input("Peer tickers (comma-separated)", value="MSFT,GOOGL,META,AMZN")
 
    # ── Overrides ─────────────────────────────────────────────
    with st.expander("Balance Sheet Overrides", expanded=False):
        override_nd = st.checkbox("Override Net Debt")        
        nd_override = st.number_input("Net Debt ($B)", value=0.0, step=0.5)*1e9 if override_nd else None
        override_sh = st.checkbox("Override Shares")
        sh_override = st.number_input("Shares (M)", value=1000.0, step=10.0)*1e6 if override_sh else None
 
    st.markdown("<div style='margin:14px 0 4px;'></div>", unsafe_allow_html=True)
    run = st.button("▶  Run Valuation", width="stretch", type="primary")
 
    # Home button
    st.markdown("<div style='margin:8px 0 0;'></div>", unsafe_allow_html=True)
    if st.button("← Home", width="stretch"):
        st.switch_page("Home.py")

# ═══════════════════════════════════════════
# MAIN AREA
# ═══════════════════════════════════════════

if not run:
    st.markdown("""
    <div style='display:flex;flex-direction:column;align-items:center;justify-content:center;
         height:70vh;text-align:center;'>
      <div style='font-family:"IBM Plex Mono",monospace;font-size:48px;color:#21262d;
           margin-bottom:16px;'>📊</div>
      <div style='font-family:"IBM Plex Mono",monospace;font-size:28px;font-weight:600;
           color:#e6edf3;letter-spacing:-1px;margin-bottom:8px;'>
        DCF Valuation Engine
      </div>
      <div style='color:#7d8590;font-size:14px;max-width:480px;line-height:1.7;margin-bottom:24px;'>
        Professional discounted cash flow model with real-time market data,
        full UFCF transparency, dual terminal value methods, and sensitivity analysis.
      </div>
      <div style='font-family:"IBM Plex Mono",monospace;font-size:11px;color:#30363d;
           text-transform:uppercase;letter-spacing:2px;'>
        Enter a ticker → Configure assumptions → Run Valuation
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Fetch data ────────────────────────────
with st.spinner(f"Fetching live data for **{ticker_input}**…"):
    data, err = fetch_data(ticker_input)

if err:
    st.error(f"❌ {err}")
    st.stop()

info      = data["info"]
company   = info.get("longName", ticker_input)
sector    = info.get("sector", "N/A")
industry  = info.get("industry", "N/A")
cur_price = info.get("currentPrice") or info.get("regularMarketPrice") or info.get("previousClose") or 0.0
mkt_cap   = info.get("marketCap", 0) or 0
shares_info = info.get("sharesOutstanding", 0) or 0

# Balance sheet items
try:
    bal = data["balance"]
    total_debt_raw = float(safe_get(bal, "Total Debt", "Long Term Debt And Capital Lease Obligation",
                                     "Long Term Debt").iloc[0])
    cash_raw = float(safe_get(bal, "Cash And Cash Equivalents",
                               "Cash Cash Equivalents And Short Term Investments",
                               "Cash And Short Term Investments").iloc[0])
    net_debt_raw = total_debt_raw - cash_raw
except Exception:
    net_debt_raw = mkt_cap * 0.1
    total_debt_raw = net_debt_raw
    cash_raw = 0

net_debt = nd_override   if (override_nd and 'nd_override' in dir()) else net_debt_raw
shares   = sh_override   if (override_sh and 'sh_override' in dir()) else shares_info
shares   = max(shares, 1e6)   # guard

# ── WACC ──────────────────────────────────
if manual_wacc:
    wacc = wacc_man
    cod, ew, dw = 0.05, 0.8, 0.2
    cost_of_equity = wacc
else:
    beta_val = beta if (beta_ov and 'beta' in dir() and beta is not None) else (info.get("beta") or 1.0)
    beta_val = float(beta_val) if beta_val else 1.0
    cost_of_equity = rf_rate + beta_val * erp
    wacc, cod, ew, dw = build_wacc(data, cost_of_equity)

# ── Historical UFCF ───────────────────────
hist_df = build_historical_ufcf(data)

# ── Projected UFCF ───────────────────────
proj_df = project_ufcf(hist_df, rev_growths, ebit_margin, da_pct, capex_pct, nwc_pct, tax_rate, n_years)

# ── DCF Valuation ─────────────────────────
val = dcf_valuation(proj_df, wacc, tgr, exit_mult, net_debt, shares)

# ── Sensitivity Tables ────────────────────
with st.spinner("Computing sensitivity tables…"):
    sens_ggm = sensitivity_table(proj_df, wacc, tgr, exit_mult, net_debt, shares, method="ggm")
    sens_em  = sensitivity_table(proj_df, wacc, tgr, exit_mult, net_debt, shares, method="em")


# ═══════════════════════════════════════════
# RENDER UI
# ═══════════════════════════════════════════

# ── Header ───────────────────────────────
col_head, col_badge = st.columns([3, 1])
with col_head:
    upside_ggm = (val["price_ggm"] / cur_price - 1) if cur_price > 0 else 0
    upside_em  = (val["price_em"]  / cur_price - 1) if cur_price > 0 else 0
    avg_upside = (upside_ggm + upside_em) / 2

    badge_cls  = "badge-bull" if avg_upside >= 0 else "badge-bear"
    badge_sign = "▲" if avg_upside >= 0 else "▼"
    st.markdown(f"""
    <div class='ticker-header'>{ticker_input}</div>
    <div class='company-name'>{company} &nbsp;·&nbsp; {sector} &nbsp;·&nbsp; {industry}</div>
    <div style='margin-top:8px;'>
      <span class='{badge_cls}'>{badge_sign} {avg_upside:+.1%} avg. upside vs current</span>
      &nbsp;
      <span style='font-family:"IBM Plex Mono",monospace;font-size:12px;color:#7d8590;'>
        Current: ${cur_price:.2f} &nbsp;·&nbsp; Mkt Cap: {fmt(mkt_cap, "$B")}
      </span>
    </div>
    """, unsafe_allow_html=True)

st.markdown('<hr class="divider">', unsafe_allow_html=True)

# ── KPI Row ───────────────────────────────
k1, k2, k3, k4, k5, k6 = st.columns(6)

def metric_card(label, value, sub="", color=""):
    return f"""
    <div class='metric-card {color}'>
      <div class='metric-label'>{label}</div>
      <div class='metric-value'>{value}</div>
      <div class='metric-sub'>{sub}</div>
    </div>"""

with k1:
    st.markdown(metric_card("GGM Price", f"${val['price_ggm']:.2f}",
                f"{upside_ggm:+.1%} vs mkt", ""), unsafe_allow_html=True)
with k2:
    st.markdown(metric_card("Exit Mult Price", f"${val['price_em']:.2f}",
                f"{upside_em:+.1%} vs mkt", "blue"), unsafe_allow_html=True)
with k3:
    st.markdown(metric_card("EV (GGM)", fmt(val["ev_ggm"],"$B"),
                f"PV FCF: {fmt(val['sum_pv_fcf'],'$B')}", "orange"), unsafe_allow_html=True)
with k4:
    st.markdown(metric_card("WACC", f"{wacc:.2%}",
                f"Ke: {cost_of_equity:.2%} | Kd: {cod:.2%}", "purple"), unsafe_allow_html=True)
with k5:
    st.markdown(metric_card("Terminal Growth", f"{tgr:.2%}",
                f"TV/EV (GGM): {val['pv_tv_ggm']/val['ev_ggm']:.0%}" if val['ev_ggm'] > 0 else "", ""), unsafe_allow_html=True)
with k6:
    st.markdown(metric_card("Net Debt", fmt(net_debt,"$B"),
                f"Shares: {shares/1e6:.0f}M", "red"), unsafe_allow_html=True)


# ── Main Tabs ─────────────────────────────
tabs = st.tabs([
    "📈  Price History",
    "🏗️  UFCF Build-up",
    "🔮  Projections",
    "⚖️  Valuation Bridge",
    "🎯  Sensitivity Analysis",
    "📊  WACC Detail",
])

# ─ Tab 0: Price History ─────────────────
with tabs[0]:
    if not data["history"].empty:
        fig_hist = plot_price_history_with_range(
            data["history"], val["price_ggm"], val["price_em"], cur_price)
        st.plotly_chart(fig_hist, width="stretch", config=CHART_CONFIG)
    else:
        st.info("No price history available.")

    st.markdown('<div class="section-header">Key Market Statistics</div>', unsafe_allow_html=True)
    ms1, ms2, ms3, ms4 = st.columns(4)
    with ms1:
        st.metric("52W High", f"${info.get('fiftyTwoWeekHigh', 0):.2f}")
        st.metric("P/E Ratio", f"{info.get('trailingPE', 0):.1f}x")
    with ms2:
        st.metric("52W Low", f"${info.get('fiftyTwoWeekLow', 0):.2f}")
        st.metric("Forward P/E", f"{info.get('forwardPE', 0):.1f}x")
    with ms3:
        st.metric("Div Yield", f"{(info.get('dividendYield', 0) or 0)*100:.2f}%")
        st.metric("EV/EBITDA", f"{info.get('enterpriseToEbitda', 0):.1f}x")
    with ms4:
        st.metric("Beta", f"{info.get('beta', 1.0):.2f}")
        st.metric("EV/Revenue", f"{info.get('enterpriseToRevenue', 0):.1f}x")


# ─ Tab 1: UFCF Build-up ─────────────────
with tabs[1]:
    st.markdown('<div class="section-header">Historical UFCF Build-up ($M)</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class='info-box'>
    UFCF = EBIT × (1 − Tax Rate) + D&A + Δ Working Capital − CapEx
    <br>This is <strong>unlevered</strong> (pre-financing) free cash flow — independent of capital structure.
    </div>
    """, unsafe_allow_html=True)

    if not hist_df.empty:
        # Render styled table
        def style_val(v, is_pct=False):
            if pd.isna(v): return "—"
            if is_pct: return f"{v:.1%}"
            return f"{v/1e6:,.1f}"

        hist_display = hist_df.copy()
        pct_rows = ["Eff. Tax Rate"]
        rows_html = ""
        for row_name in hist_display.index:
            is_pct = row_name in pct_rows
            total_cls = "row-total" if row_name in ["UFCF", "NOPAT"] else ""
            cells = "".join(f"<td>{style_val(v, is_pct)}</td>"
                            for v in hist_display.loc[row_name])
            rows_html += f"<tr class='{total_cls}'><td>{row_name}</td>{cells}</tr>"

        header_cells = "".join(f"<th>{c}</th>" for c in hist_display.columns)
        st.markdown(f"""
        <table class='styled-table'>
          <thead><tr><th>Line Item</th>{header_cells}</tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
        """, unsafe_allow_html=True)

        st.plotly_chart(plot_ufcf_history(hist_df), width="stretch", config=CHART_CONFIG)
    else:
        st.warning("Could not compute historical UFCF — check ticker or try a different company.")


# ─ Tab 2: Projections ───────────────────
with tabs[2]:
    st.markdown('<div class="section-header">Projected UFCF Build-up ($M)</div>', unsafe_allow_html=True)

    # Render projection table
    proj_display = proj_df.copy()
    dollar_c = ["Revenue", "EBIT", "NOPAT", "D&A", "ΔWorking Cap", "CapEx", "UFCF"]
    pct_c    = ["Rev Growth", "EBIT Margin", "Tax Rate", "D&A % Rev", "NWC % Rev", "CapEx % Rev"]

    def fmt_proj(col_name, v):
        if pd.isna(v): return "—"
        if col_name in pct_c: return f"{v:.1%}"
        return f"{v/1e6:,.1f}"

    years_list = proj_df.index.tolist()
    header_cells = "".join(f"<th>{y}</th>" for y in years_list)
    rows_html = ""
    row_order = ["Revenue","Rev Growth","EBIT","EBIT Margin","Tax Rate",
                 "NOPAT","D&A","D&A % Rev","ΔWorking Cap","NWC % Rev","CapEx","CapEx % Rev","UFCF"]
    for rn in row_order:
        if rn not in proj_display.columns: continue
        total_cls = "row-total" if rn in ["UFCF","NOPAT"] else ""
        cells = "".join(f"<td>{fmt_proj(rn, proj_display.loc[y, rn])}</td>" for y in years_list)
        rows_html += f"<tr class='{total_cls}'><td>{rn}</td>{cells}</tr>"

    st.markdown(f"""
    <table class='styled-table'>
      <thead><tr><th>Line Item</th>{header_cells}</tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    """, unsafe_allow_html=True)

    st.plotly_chart(plot_projected_fcf(proj_df, val["pv_fcf"]), width="stretch", config=CHART_CONFIG)


# ─ Tab 3: Valuation Bridge ──────────────
with tabs[3]:
    st.markdown('<div class="section-header">Valuation Bridge & Summary</div>', unsafe_allow_html=True)

    left, right = st.columns(2)

    with left:
        st.markdown("#### Gordon Growth Model")
        bridge_ggm = {
            "PV of Projected FCFs":        val["sum_pv_fcf"],
            "Terminal Value (GGM)":         val["tv_ggm"],
            "PV of Terminal Value":         val["pv_tv_ggm"],
            "Enterprise Value":             val["ev_ggm"],
            "Less: Net Debt":              -net_debt,
            "Equity Value":                 val["eq_ggm"],
            "Shares Outstanding (M)":       shares / 1e6,
            "Implied Share Price":          val["price_ggm"],
        }
        for k, v in bridge_ggm.items():
            sep = "row-total" if k in ["Enterprise Value", "Implied Share Price"] else ""
            if k == "Shares Outstanding (M)":
                st.markdown(f"**{k}**: `{v:,.1f}M`")
            elif k == "Implied Share Price":
                st.markdown(f"**{k}**: `${v:.2f}`")
            elif abs(v) > 1e8:
                st.markdown(f"**{k}**: `{fmt(v, '$B')}`")
            else:
                st.markdown(f"**{k}**: `{v:.2f}`")

        tv_pct = val["pv_tv_ggm"] / val["ev_ggm"] * 100 if val["ev_ggm"] > 0 else 0
        st.markdown(f"*TV represents **{tv_pct:.0f}%** of EV — {'⚠️ high sensitivity to terminal assumptions' if tv_pct > 75 else '✅ reasonable terminal value weight'}*")
        st.plotly_chart(plot_valuation_bridge(val, net_debt, shares, method="ggm"), width="stretch", config=CHART_CONFIG)

    with right:
        st.markdown("#### Exit Multiple Method")
        bridge_em = {
            "PV of Projected FCFs":        val["sum_pv_fcf"],
            "Terminal EBIT":               proj_df["EBIT"].iloc[-1],
            f"Exit EV/EBITDA ({exit_mult:.1f}x)": val["tv_em"],
            "PV of Terminal Value":         val["pv_tv_em"],
            "Enterprise Value":             val["ev_em"],
            "Less: Net Debt":              -net_debt,
            "Equity Value":                 val["eq_em"],
            "Shares Outstanding (M)":       shares / 1e6,
            "Implied Share Price":          val["price_em"],
        }
        for k, v in bridge_em.items():
            if k == "Shares Outstanding (M)":
                st.markdown(f"**{k}**: `{v:,.1f}M`")
            elif k == "Implied Share Price":
                st.markdown(f"**{k}**: `${v:.2f}`")
            elif abs(v) > 1e8:
                st.markdown(f"**{k}**: `{fmt(v, '$B')}`")
            else:
                st.markdown(f"**{k}**: `{v:.2f}`")

        tv_pct_em = val["pv_tv_em"] / val["ev_em"] * 100 if val["ev_em"] > 0 else 0
        st.markdown(f"*TV represents **{tv_pct_em:.0f}%** of EV — {'⚠️ high sensitivity to exit multiple' if tv_pct_em > 75 else '✅ reasonable terminal value weight'}*")
        st.plotly_chart(plot_valuation_bridge(val, net_debt, shares, method="em"), width="stretch", config=CHART_CONFIG)

    # PV of each year's FCF
    st.markdown('<div class="section-header">Annual PV of FCF ($M)</div>', unsafe_allow_html=True)
    years_list = proj_df.index.tolist()
    fcf_vals   = proj_df["UFCF"].values / 1e6
    pv_vals    = val["pv_fcf"] / 1e6

    fig_pv = go.Figure()
    fig_pv.add_trace(go.Bar(x=years_list, y=fcf_vals, name="Nominal UFCF",
                            marker_color="#388bfd", opacity=0.5))
    fig_pv.add_trace(go.Bar(x=years_list, y=pv_vals, name="PV of UFCF",
                            marker_color="#3fb950"))
    fig_pv.update_layout(barmode="group", title="Nominal vs. Present Value of FCF ($M)",
                         **CHART_LAYOUT)
    st.plotly_chart(fig_pv, width="stretch", config=CHART_CONFIG)


# ─ Tab 4: Sensitivity ───────────────────
with tabs[4]:
    st.markdown('<div class="section-header">Sensitivity Analysis — Implied Share Price</div>',
                unsafe_allow_html=True)

    st.markdown("""
    <div class='info-box'>
    Each cell shows the implied share price for a given WACC × terminal assumption combination.
    Green = above current market price. Red = below market price.
    </div>
    """, unsafe_allow_html=True)

    sA, sB = st.columns(2)
    with sA:
        st.markdown("#### GGM — WACC × Terminal Growth Rate")
        st.plotly_chart(plot_sensitivity_heatmap(sens_ggm, cur_price), width="stretch", config=CHART_CONFIG)
        st.dataframe(sens_ggm.style.background_gradient(cmap="RdYlGn", axis=None)
                     .format("${:.2f}"), width="stretch")

    with sB:
        st.markdown("#### Exit Multiple — WACC × EV/EBITDA")
        st.plotly_chart(plot_sensitivity_heatmap(sens_em, cur_price), width="stretch", config=CHART_CONFIG)
        st.dataframe(sens_em.style.background_gradient(cmap="RdYlGn", axis=None)
                     .format("${:.2f}"), width="stretch")

    # Tornado chart — WACC vs TGR impact
    st.markdown('<div class="section-header">Tornado: Key Assumption Impact</div>',
                unsafe_allow_html=True)

    base_price = (val["price_ggm"] + val["price_em"]) / 2

    def tornado_var(param, low_val, high_val, **kwargs):
        """Returns (low_price, high_price, param_name)."""
        base = dict(proj_df=proj_df, wacc=wacc, tgr=tgr, exit_ebitda_mult=exit_mult,
                    net_debt=net_debt, shares_out=shares)
        kw_low  = {**base, param: low_val}
        kw_high = {**base, param: high_val}
        vl = dcf_valuation(**kw_low)
        vh = dcf_valuation(**kw_high)
        pl = (vl["price_ggm"] + vl["price_em"]) / 2
        ph = (vh["price_ggm"] + vh["price_em"]) / 2
        return pl, ph

    params_tornado = [
        ("WACC",          "wacc",            wacc - 0.02,   wacc + 0.02),
        ("Terminal Growth","tgr",            tgr  - 0.01,   tgr  + 0.01),
        ("Exit Multiple",  "exit_ebitda_mult",exit_mult-3,  exit_mult+3),
    ]

    names, lows, highs = [], [], []
    for label, param, lo, hi in params_tornado:
        pl, ph = tornado_var(param, lo, hi)
        names.append(label)
        lows.append(min(pl, ph))
        highs.append(max(pl, ph))

    fig_torn = go.Figure()
    for i, (n, l, h) in enumerate(zip(names, lows, highs)):
        fig_torn.add_trace(go.Bar(
            y=[n], x=[l - base_price], base=[base_price],
            orientation="h", name="Low",
            marker_color="#f85149", showlegend=(i==0)
        ))
        fig_torn.add_trace(go.Bar(
            y=[n], x=[h - base_price], base=[base_price],
            orientation="h", name="High",
            marker_color="#3fb950", showlegend=(i==0)
        ))
    fig_torn.add_vline(x=base_price, line_dash="dash", line_color="#7d8590",
                       annotation_text=f"Base: ${base_price:.2f}")
    fig_torn.update_layout(title="Tornado: Assumption Sensitivity (Avg Price $)",
                           barmode="overlay", **CHART_LAYOUT, height=280)
    st.plotly_chart(fig_torn, width="stretch", config=CHART_CONFIG)


# ─ Tab 5: WACC Detail ────────────────────
with tabs[5]:
    st.markdown('<div class="section-header">WACC Decomposition</div>', unsafe_allow_html=True)

    beta_display = info.get("beta") or 1.0
    if 'beta_ov' in dir() and beta_ov and beta is not None:
        beta_display = beta

    w1, w2 = st.columns(2)
    with w1:
        st.markdown("#### Cost of Equity (CAPM)")
        capm_items = {
            "Risk-Free Rate (Rf)":         f"{rf_rate:.2%}" if not manual_wacc else "—",
            "Beta (β)":                    f"{beta_display:.2f}",
            "Equity Risk Premium (ERP)":   f"{erp:.2%}" if not manual_wacc else "—",
            "Cost of Equity (Ke)":         f"{cost_of_equity:.2%}",
            "Equity Weight (E/V)":         f"{ew:.1%}",
        }
        for k, v in capm_items.items():
            col_a, col_b = st.columns([2, 1])
            col_a.write(k)
            col_b.markdown(f"`{v}`")

    with w2:
        st.markdown("#### Cost of Debt & WACC")
        wacc_items = {
            "Cost of Debt (Pre-tax Kd)":   f"{cod:.2%}",
            "Tax Rate":                    f"{tax_rate:.1%}",
            "After-tax Kd":                f"{cod*(1-tax_rate):.2%}",
            "Debt Weight (D/V)":           f"{dw:.1%}",
            "─────────────────────":       "─────────",
            "WACC":                        f"{wacc:.2%}",
        }
        for k, v in wacc_items.items():
            col_a, col_b = st.columns([2, 1])
            col_a.write(k)
            col_b.markdown(f"`{v}`")

    # WACC waterfall
    fig_wacc = go.Figure(go.Bar(
        x=["Ke × Equity Wt", "Kd(AT) × Debt Wt", "WACC"],
        y=[cost_of_equity * ew, cod * (1 - tax_rate) * dw, wacc],
        marker_color=["#1f6feb", "#d29922", "#3fb950"],
        text=[f"{cost_of_equity*ew:.2%}", f"{cod*(1-tax_rate)*dw:.2%}", f"{wacc:.2%}"],
        textposition="outside",
    ))
    fig_wacc.update_layout(title="WACC Composition", **CHART_LAYOUT, height=320)
    st.plotly_chart(fig_wacc, width="stretch", config=CHART_CONFIG)


# ═══════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════
st.markdown('<hr class="divider">', unsafe_allow_html=True)

st.markdown('<div class="section-header">Export</div>', unsafe_allow_html=True)

ec1, ec2, ec3 = st.columns([2, 2, 3])
with ec1:
    try:
        xl_bytes = export_excel(
            ticker_input, info, hist_df, proj_df, val, sens_ggm, sens_em,
            wacc, tgr, exit_mult, net_debt, shares, cur_price
        )
        st.download_button(
            label="⬇  Download Full Excel Model",
            data=xl_bytes,
            file_name=f"{ticker_input}_DCF_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            type="primary",
        )
    except Exception as ex:
        st.error(f"Excel export failed: {ex}")

with ec2:
    # Quick CSV of projection
    csv_buf = proj_df.to_csv().encode()
    st.download_button(
        label="⬇  Download Projections (CSV)",
        data=csv_buf,
        file_name=f"{ticker_input}_projections.csv",
        mime="text/csv",
        width="stretch",
    )

with ec3:
    st.markdown(f"""
    <div style='font-family:"IBM Plex Mono",monospace;font-size:11px;color:#7d8590;padding:8px 0;'>
    📦 Excel includes: Summary · Historical UFCF · Projected UFCF · DCF Bridge ·
    Sensitivity (GGM) · Sensitivity (Exit Multiple) — formatted with styled headers.
    </div>
    """, unsafe_allow_html=True)

st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
st.markdown("""
<div style='font-family:"IBM Plex Mono",monospace;font-size:10px;color:#30363d;
     text-align:center;padding:16px 0;'>
DCF Engine — Data via yfinance · For informational purposes only · Not financial advice
</div>
""", unsafe_allow_html=True)
